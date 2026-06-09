import os
import re
from collections import Counter, defaultdict
from statistics import median
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd

BASE_DIR = Path(os.getenv("QUOTE_TOOL_DIR", r"C:\Quote_Tool"))
DATA_DIR = BASE_DIR / "DATA"

# If the script is run from a copied folder and C:\Quote_Tool\DATA does not exist,
# fall back to a sibling DATA folder. This makes the tool easier to move to a new PC.
SCRIPT_DIR = Path(__file__).resolve().parent
if not DATA_DIR.exists():
    sibling_data = SCRIPT_DIR.parent / "DATA"
    local_data = SCRIPT_DIR / "DATA"
    if sibling_data.exists():
        BASE_DIR = SCRIPT_DIR.parent
        DATA_DIR = sibling_data
    elif local_data.exists():
        BASE_DIR = SCRIPT_DIR
        DATA_DIR = local_data

EXCEL_EXTENSIONS = (".xlsx", ".xls", ".xlsm", ".xlsb")


def file_score_for_name(path, preferred_names):
    name = path.name.lower()
    compact_name = re.sub(r"[^a-z0-9]", "", name)
    best = 0
    for pref in preferred_names:
        pref_l = pref.lower()
        pref_compact = re.sub(r"[^a-z0-9]", "", pref_l)
        if name == pref_l:
            best = max(best, 100)
        elif name.startswith(pref_l):
            best = max(best, 90)
        elif compact_name.startswith(pref_compact):
            best = max(best, 80)
        elif pref_compact in compact_name:
            best = max(best, 55)
    return best


def find_data_file(preferred_names, extensions=EXCEL_EXTENSIONS):
    """Find a DATA file even if Windows added double extensions.

    Examples this handles:
      incoming_quotes.xlsx
      incoming_quotes.xlsx.xlsx
      purchase_orders.xlsx.xls
      9781-2_ILS.txt.txt
    """
    if not DATA_DIR.exists():
        return None

    candidates = []
    for path in DATA_DIR.iterdir():
        if not path.is_file():
            continue
        lower = path.name.lower()
        if extensions and not any(lower.endswith(ext.lower()) for ext in extensions):
            continue
        score = file_score_for_name(path, preferred_names)
        if score > 0:
            candidates.append((score, path.stat().st_mtime, path))

    if not candidates:
        return None

    candidates.sort(key=lambda x: (x[0], x[1]), reverse=True)
    return candidates[0][2]


OUTGOING_FILE_CANDIDATES = ["outgoing_quotes.xlsx", "outgoing_quotes", "outgoing", "quotes_out", "sell_quotes"]
INCOMING_FILE_CANDIDATES = ["incoming_quotes.xlsx", "incoming_quotes", "incoming", "quotes_in", "vendor_quotes"]
PO_FILE_CANDIDATES = ["purchase_orders.xlsx", "purchase_orders", "po_history", "po", "purchases"]
SALES_HISTORY_CANDIDATES = [
    "sales_orders.xlsx", "sales_orders", "sale_orders", "sales_order", "sales",
    "so", "sos", "sales_history", "completed_sales", "invoices", "invoice_history"
]
INVENTORY_FILE_CANDIDATES = [
    "inventory.xlsx", "inventory", "inventory_report", "all_inventory", "stock", "on_hand", "onhand"
]

ILS_FILE = DATA_DIR / "Print_Wizard_SearchResults_05_08_26.txt"  # fallback only

TODAY = datetime.today()

COND_MAP = {
    "overhauled": "OH", "oh": "OH",
    "repaired": "RP", "rp": "RP",
    "inspected": "IN", "inspected/tested": "IN", "in": "IN",
    "serviceable": "SV", "sv": "SV",
    "new_item": "NE", "new item": "NE", "new": "NE", "ne": "NE",
    "as_removed": "AR", "as removed": "AR", "ar": "AR",
    "ns": "NS",
}

VISIBLE_BUY_CONDITIONS = ["OH", "RP", "IN", "SV", "NE"]
MIN_SELL_MARKUP_OVER_BUY_MAX = 1.30

PART_FIELD_NAMES = [
    "PART #", "Part #", "Part Number", "PART NUMBER", "P/N", "PN", "P N",
    "Part No", "PART NO", "Item", "Item Number", "Item #", "ITEM #",
    "Product", "Product Code", "Part", "Component"
]

COND_FIELD_NAMES = [
    "Condition", "CONDITION", "Cond", "COND", "Part Condition", "CD", "Code"
]

BUY_PRICE_FIELD_NAMES = [
    "EXCH FEE / COST EA", "EXCH. FEE", "Exchange Fee / Cost EA",
    "COST PER UNIT", "Cost Per Unit", "Unit Cost", "Cost", "Purchase Price",
    "PO Price", "Price Paid", "Unit Price", "Unit Amount", "OUTRIGHT VALUE",
    "Outright Value", "Price", "Amount", "Ext Cost", "Extended Cost", "TOTAL COST", "Total Cost"
]

SELL_PRICE_FIELD_NAMES = [
    "EXCH FEE / SALES EA", "Exchange Fee / Sales EA", "OUTRIGHT VALUE", "Outright Value",
    "Price Per Unit", "PRICE PER UNIT", "Unit Price", "Unit Amount", "Sell Price",
    "Sale Price", "Sales Price", "SO Price", "S/O Price", "Invoice Price",
    "Sold Price", "Sold For", "Unit Sell", "Sell", "Sales Amount", "Invoice Amount",
    "AMOUNT PAID", "Amount Paid", "Price", "Amount", "Revenue", "Ext Price", "Extended Price", "Line Total",
    "Total", "Net", "Net Price"
]

DATE_FIELD_NAMES = [
    "Date", "DATE", "Quote Date", "QUOTE DATE", "Created Date", "CREATED DATE",
    "CREATED/UPDATED", "Created/Updated", "DATE CREATED", "Date Created", "DATE RECEIVED", "Date Received",
    "PO Date", "P.O. Date", "Purchase Date", "Order Date", "SO Date", "S/O Date",
    "Sales Date", "Invoice Date", "DATE SHIPPED", "Date Shipped", "Shipped Date", "Closed Date", "Entry Date"
]

INVENTORY_COST_FIELD_NAMES = [
    "TOTAL COST", "Total Cost", "COST/UNIT", "Cost/Unit", "COST PER UNIT", "Cost Per Unit", "Unit Cost", "Cost", "Average Cost"
]

INVENTORY_SRP_FIELD_NAMES = [
    "SRP", "Suggested Price", "List Price", "Target Sell"
]

INVENTORY_EXCH_SRP_FIELD_NAMES = [
    "EXCH SRP", "Exchange SRP", "Exch SRP"
]

INVENTORY_SALE_PRICE_FIELD_NAMES = [
    "SALES PRICE", "Sales Price", "SALE PRICE", "Sale Price", "Sold Price", "Sold For"
]

INVENTORY_LOCATION_FIELD_NAMES = [
    "INV LOCATION", "Inventory Location", "Location", "LOCATION", "WH LOCATION", "Warehouse Location", "STORE LOCATION"
]

INVENTORY_VENDOR_FIELD_NAMES = [
    "PURCHASED FROM", "Purchased From", "Vendor", "VENDOR", "Supplier", "Purchased Vendor"
]

INVENTORY_TAG_BY_FIELD_NAMES = [
    "TAGGED BY", "Tagged By", "Tag By"
]

INVENTORY_TAG_DATE_FIELD_NAMES = [
    "TAG DATE", "Tag Date", "Tag", "Cert Date", "Certification Date"
]

# Keep for backward compat in read_excel_file column selection
INVENTORY_QTY_FIELD_NAMES = [
    "QTY", "Qty", "Quantity", "QUANTITY", "On Hand", "ON HAND", "Available Qty", "Available"
]


EXTRA_RELEVANT_FIELD_NAMES = [
    "VENDOR", "Vendor", "Vendor Name", "Company", "Supplier", "Customer", "CUSTOMER",
    "CANCELLED", "Cancelled", "CANCELED", "Canceled", "Cancellation Reason", "CANCELLATION REASON",
    "SOURCE", "Source", "FILE", "File", "FILE NAME", "File Name", "MODULE", "Module",
    "DOC TYPE", "Doc Type", "DOCUMENT TYPE", "Transaction Type", "TRANSACTION TYPE", "TYPE", "Type",
    "SALES PRICE", "Sales Price", "COST AT TIME OF SALE", "Cost At Time Of Sale",
]

RELEVANT_EXCEL_FIELD_NAMES = list(dict.fromkeys(
    PART_FIELD_NAMES + COND_FIELD_NAMES + BUY_PRICE_FIELD_NAMES + SELL_PRICE_FIELD_NAMES + DATE_FIELD_NAMES +
    INVENTORY_QTY_FIELD_NAMES + INVENTORY_COST_FIELD_NAMES + INVENTORY_SRP_FIELD_NAMES +
    INVENTORY_LOCATION_FIELD_NAMES + INVENTORY_VENDOR_FIELD_NAMES + INVENTORY_TAG_DATE_FIELD_NAMES +
    INVENTORY_EXCH_SRP_FIELD_NAMES + INVENTORY_SALE_PRICE_FIELD_NAMES + INVENTORY_TAG_BY_FIELD_NAMES +
    EXTRA_RELEVANT_FIELD_NAMES
))

SOURCE_WEIGHT = {
    "completed PO history": 1.35,
    "completed sales history": 1.45,
    "incoming quotes": 0.75,
    "outgoing quotes": 0.75,
}

CONDITION_SANITY_CAPS = {
    # These conditions should not price higher than OH in normal outright guidance.
    "RP": 0.85,
    "IN": 0.75,
    "SV": 0.70,
}


COND_RANK = {
    "NE": 6,
    "OH": 5,
    "RP": 4,
    "SV": 3,
    "IN": 2,
    "NS": 1,
    "AR": 0,
}

SECTION_RANK = {
    "daily": 4,
    "weekly": 3,
    "monthly": 2,
    "old": 1,
    "unknown": 0,
}

VENDOR_ALIASES = {
    "KILLICK AEROSPACE LTD": "KILLICK AEROSPACE",
    "KILLICK AEROSPACE INC": "KILLICK AEROSPACE",
    "SENTRY AEROSPARES LTD": "SENTRY AEROSPARES",
    "SENTRY AEROSPARES LLC": "SENTRY AEROSPARES",
    "MAGELLAN GROUP SHANNON": "MAGELLAN AVIATION GROUP",
    "MAGELLAN AVIATION GROUP": "MAGELLAN AVIATION GROUP",
    "AAR AIRCRAFT TURBINE CENTER": "AAR",
    "AAR ALLEN ASSET MANAGEMENT": "AAR",
    "AAR A/C COMPONENT SVCS - NY": "AAR",
    "AAR COMPONENT SERVICES - GRAND PRAIRIE": "AAR",

    # Company family / preferred internal relationship vendors.
    # These normalize common ways the companies may appear in ILS or PO history.
    "SUPPORT AIR INC": "SUPPORT AIR",
    "SUPPORT AIR": "SUPPORT AIR",

    "BROWARD AVIATION COMPANY": "BROWARD AVIATION COMPANY",
    "BROWARD AVIATION CO": "BROWARD AVIATION COMPANY",
    "BROWARD AVIATION SERVICES INC": "BROWARD AVIATION COMPANY",
    "BROWARD AVIATION SERVICES": "BROWARD AVIATION COMPANY",

    "AIR ACCESSORIES AND AVIONICS": "AIR ACCESSORIES AND AVIONICS",
    "AIR ACCESSORIES & AVIONICS": "AIR ACCESSORIES AND AVIONICS",
    "AIR ACCESSORIES AVIONICS": "AIR ACCESSORIES AND AVIONICS",

    "JET AIR MRO": "JET AIR MRO",
    "JET AIR MRO LLC": "JET AIR MRO",
}

OUR_COMPANY_NAMES = {"SUPPORT AIR"}
SISTER_COMPANY_NAMES = {
    "BROWARD AVIATION COMPANY",
    "AIR ACCESSORIES AND AVIONICS",
    "JET AIR MRO",
}

def company_relationship(vendor_name):
    """Return preferred-company relationship tier for call-list sorting.

    0 = our own company
    1 = sister company / preferred affiliated company
    2 = normal outside vendor
    """
    vendor = normalize_vendor(vendor_name)
    if vendor in OUR_COMPANY_NAMES:
        return 0, "own company"
    if vendor in SISTER_COMPANY_NAMES:
        return 1, "sister company"
    return 2, ""

VENDOR_LINE_RE = re.compile(r"^\s*(?:\+?\d|1-|44|353|31|45|49|65|90|971|48|852|254)")
PART_LINE_RE = re.compile(
    # ILS availability line. Handles both blank CAGE and populated CAGE.
    # Examples:
    #      01   9781-2               OH           2 LATCH
    #      01   3202222-1            OH  7BP93    1 BLEED AIR CHECK VALV
    # M    01   9781-2               NE        RQST LATCH ASSY
    r"^\s*M?\s*\d{2}\s+([A-Z0-9\-]+)\s+([A-Z]{2})\s+(?:(?!\d+\b|RQST\b)([A-Z0-9\-]+)\s+)?(\d+|RQST)\s*(.*)$"
)

def clean(value):
    if value is None:
        return ""
    return str(value).strip()

def clean_cond(value):
    value = clean(value).lower()
    return COND_MAP.get(value, value.upper())

def normalize_vendor(vendor):
    vendor = clean(vendor).upper()
    vendor = vendor.replace(".", "").replace(",", "")
    vendor = " ".join(vendor.split())
    return VENDOR_ALIASES.get(vendor, vendor)

def money(value):
    try:
        if pd.isna(value):
            return None
        return float(str(value).replace("$", "").replace(",", "").strip())
    except:
        return None

def fmt_money(value):
    return f"${value:,.2f}"

def parse_date(value):
    if value is None or clean(value) == "":
        return None
    try:
        dt = pd.to_datetime(value, errors="coerce")
        if pd.isna(dt):
            return None
        return dt.to_pydatetime()
    except:
        return None

def compact_field_name(value):
    return re.sub(r"[^a-z0-9]", "", str(value).lower())


def select_relevant_columns(columns, wanted_names=RELEVANT_EXCEL_FIELD_NAMES):
    wanted = [compact_field_name(n) for n in wanted_names]
    selected = []
    for col in columns:
        cc = compact_field_name(col)
        if not cc:
            continue
        if cc in wanted or any(w and (w in cc or cc in w) for w in wanted):
            selected.append(col)
    return selected or list(columns)


def read_excel_file(path):
    if not path or not path.exists():
        print(f"WARNING: Missing file: {path}")
        return []

    try:
        # First read headers only, then load only columns the tool actually uses.
        # This keeps large exports faster and much lighter on memory.
        header_df = pd.read_excel(path, nrows=0)
        usecols = select_relevant_columns(header_df.columns)
        df = pd.read_excel(path, usecols=usecols)
        df = df.fillna("")
        df.columns = [clean(c) for c in df.columns]
        return df.to_dict("records")
    except Exception as e:
        print(f"ERROR reading {path.name}: {e}")
        return []

def read_excel_file_for_part(path, part, part_names=PART_FIELD_NAMES):
    """Read only rows matching one part number using a memory-light streaming reader.

    This keeps the inventory report from consuming a lot of memory at startup.
    It works with normal .xlsx files and double-extension names like inventory.xlsx.xlsx.
    """
    if not path or not path.exists():
        return []

    search = part.upper().strip()

    def compact(s):
        return re.sub(r"[^a-z0-9]", "", str(s).lower())

    part_name_compacts = [compact(n) for n in part_names]

    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        headers = next(rows_iter, None)
        if not headers:
            wb.close()
            return []

        headers = [clean(h) for h in headers]
        header_compacts = [compact(h) for h in headers]
        part_idx = None
        for i, hc in enumerate(header_compacts):
            if hc in part_name_compacts or any(pc and pc in hc for pc in part_name_compacts):
                part_idx = i
                break

        if part_idx is None:
            wb.close()
            return []

        matches = []
        for values in rows_iter:
            if part_idx >= len(values):
                continue
            pn = clean(values[part_idx]).upper()
            if pn != search:
                continue
            row = {}
            for i, header in enumerate(headers):
                if header:
                    row[header] = values[i] if i < len(values) else ""
            matches.append(row)

        wb.close()
        return matches

    except Exception:
        # Fallback to pandas if openpyxl is unavailable or a workbook is unusual.
        try:
            df = pd.read_excel(path)
            df = df.fillna("")
            df.columns = [clean(c) for c in df.columns]
            part_col = None
            for c in df.columns:
                cc = compact(c)
                if cc in part_name_compacts or any(pc and pc in cc for pc in part_name_compacts):
                    part_col = c
                    break
            if part_col is None:
                return []
            mask = df[part_col].astype(str).str.strip().str.upper() == search
            return df.loc[mask].to_dict("records")
        except Exception as e:
            print(f"ERROR reading {path.name} for inventory/part lookup: {e}")
            return []

def read_first_existing_excel(names_or_paths):
    # Accept either real Path objects or preferred name strings.
    paths = [p for p in names_or_paths if isinstance(p, Path)]
    names = [str(p) for p in names_or_paths if not isinstance(p, Path)]

    for path in paths:
        if path.exists():
            return read_excel_file(path), path

    if names:
        path = find_data_file(names)
        if path:
            return read_excel_file(path), path

    return [], None

def find_latest_ils_file():
    """Use the newest .txt ILS export in DATA when available.
    This prevents the tool from saying 'No ILS file loaded' just because
    the newest Print Wizard file has a different date in the filename.
    """
    if DATA_DIR.exists():
        txt_files = [p for p in DATA_DIR.glob("*.txt") if p.is_file()]
        if txt_files:
            return max(txt_files, key=lambda p: p.stat().st_mtime)

    return ILS_FILE

def find_all_ils_files():
    """Load every ILS text export in DATA.

    This accepts both normal .txt files and double-extension files such as
    9781-2_ILS.txt.txt. It also avoids loading unrelated text files unless
    no ILS-named files are found.
    """
    if DATA_DIR.exists():
        txt_files = [p for p in DATA_DIR.iterdir() if p.is_file() and p.name.lower().endswith(".txt")]
        ils_named = [p for p in txt_files if "ils" in p.name.lower() or "print_wizard" in p.name.lower()]
        selected = ils_named if ils_named else txt_files
        if selected:
            return sorted(selected, key=lambda p: p.stat().st_mtime, reverse=True)

    return [ILS_FILE] if ILS_FILE.exists() else []

def find_value(row, names):
    # This version is intentionally very flexible:
    # it handles PART NUMBER vs Part Number vs PART #, etc.
    normalized = {}

    for key in row.keys():
        k = str(key).strip()
        normalized[k.lower()] = key
        normalized[k.upper()] = key

    for name in names:
        n1 = name.strip().lower()
        n2 = name.strip().upper()

        if n1 in normalized:
            return row.get(normalized[n1], "")
        if n2 in normalized:
            return row.get(normalized[n2], "")

    # fallback: compare names with punctuation/spaces removed
    def compact(s):
        return re.sub(r"[^a-z0-9]", "", str(s).lower())

    compact_map = {compact(k): k for k in row.keys()}

    for name in names:
        c = compact(name)
        if c in compact_map:
            return row.get(compact_map[c], "")

    # contains fallback
    for key in row.keys():
        key_c = compact(key)
        for name in names:
            name_c = compact(name)
            if name_c and name_c in key_c:
                return row.get(key, "")

    return ""


def build_part_index(rows, part_names=PART_FIELD_NAMES):
    """Build a fast part-number lookup index so searches do not scan big Excel exports every time."""
    index = defaultdict(list)
    for row in rows:
        pn = clean(find_value(row, part_names)).upper()
        if pn:
            index[pn].append(row)
    return index


def rows_for_part(rows_or_index, search):
    if isinstance(rows_or_index, dict):
        return list(rows_or_index.get(search.upper(), []))
    return rows_or_index

def detect_section(line, current):
    upper = line.upper()
    if "UPDATED NO LESS THAN DAILY" in upper:
        return "daily"
    if "UPDATED NO LESS THAN WEEKLY" in upper:
        return "weekly"
    if "UPDATED NO LESS THAN MONTHLY" in upper:
        return "monthly"
    if "NOT BEEN UPDATED IN MORE THAN 45 DAYS" in upper:
        return "old"
    return current

def parse_vendor_line(line):
    text = clean(line)
    cage_match = re.search(r"\(([A-Z0-9]{4})\)\s*$", text)
    no_cage = re.sub(r"\([A-Z0-9]{4}\)\s*$", "", text).strip()

    parts = no_cage.split()
    vendor_words = []

    for token in parts:
        if any(ch.isalpha() for ch in token):
            vendor_words.append(token)
        elif vendor_words:
            vendor_words.append(token)

    vendor = " ".join(vendor_words).strip()
    if not vendor:
        vendor = text

    return normalize_vendor(vendor)

def parse_ils_file(path):
    if not path.exists():
        return []

    vendors = []
    current_vendor = None
    current_section = "unknown"

    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        for raw_line in f:
            line = raw_line.rstrip("\n")
            current_section = detect_section(line, current_section)

            if "OVERHAUL CAPABILITY" in line.upper():
                break

            # Vendor header lines end with a 4-character ILS code in parentheses.
            # Contact lines may also contain phone numbers in parentheses, so avoid
            # accidentally treating those as new vendors.
            if VENDOR_LINE_RE.match(line) and re.search(r"\([A-Z0-9]{4}\)\s*$", line):
                vendor_name = parse_vendor_line(line)
                current_vendor = {
                    "vendor": vendor_name,
                    "section": current_section,
                    "contacts": [],
                    "listings": [],
                }
                vendors.append(current_vendor)
                continue

            if current_vendor is not None:
                upper = line.upper()

                if "EMAIL" in upper or "SALES" in upper or "AOG" in upper or "RFQ" in upper or "@" in line:
                    current_vendor["contacts"].append(clean(line))

                m = PART_LINE_RE.match(line)

                if m:
                    part, cond, cage_field, qty, desc = m.groups()
                    current_vendor["listings"].append(
                        {
                            "part": clean(part).upper(),
                            "cond": clean_cond(cond),
                            "qty": clean(qty),
                            "description": clean(desc),
                        }
                    )

    merged = {}

    for vendor in vendors:
        key = vendor["vendor"]

        if key not in merged:
            merged[key] = {
                "vendor": key,
                "sections": [],
                "contacts": [],
                "listings": [],
            }

        merged[key]["sections"].append(vendor["section"])
        merged[key]["contacts"].extend(vendor["contacts"])
        merged[key]["listings"].extend(vendor["listings"])

    output = []

    for vendor in merged.values():
        best_section = max(vendor["sections"], key=lambda s: SECTION_RANK.get(s, 0)) if vendor["sections"] else "unknown"
        output.append(
            {
                "vendor": vendor["vendor"],
                "section": best_section,
                "contacts": list(dict.fromkeys(vendor["contacts"])),
                "listings": vendor["listings"],
            }
        )

    return output

def load_po_history(rows):
    history = defaultdict(lambda: {
        "total_po_count": 0,
        "part_po_count": defaultdict(int),
    })

    for row in rows:
        if row_is_cancelled(row):
            continue
        vendor = normalize_vendor(find_value(row, ["VENDOR", "Vendor", "Vendor Name", "Company", "Supplier"]))
        part = clean(find_value(row, ["PART NUMBER", "Part Number", "Part #", "P/N", "PN"])).upper()

        if not vendor:
            continue

        history[vendor]["total_po_count"] += 1

        if part:
            history[vendor]["part_po_count"][part] += 1

    return history

def confidence(count):
    if count >= 6:
        return "HIGH"
    if count >= 3:
        return "MED"
    return "LOW"

def remove_outliers(prices):
    if len(prices) < 4:
        return prices, 0

    med = median(prices)
    filtered = []
    removed = 0

    for p in prices:
        if med * 0.5 <= p <= med * 1.5:
            filtered.append(p)
        else:
            removed += 1

    return filtered, removed

def row_text(row):
    return " ".join(clean(v) for v in row.values()).upper()

def is_repair_source_row(row):
    """Return True only when the ROW SOURCE/TYPE clearly says it came from repair data.

    We do NOT remove condition RP. RP is a valid market condition.
    We also do NOT remove descriptions that contain words like repair/overhaul,
    because that was hiding legitimate RP and IN records.
    """
    source_fields = [
        "SOURCE", "Source", "FILE", "File", "FILE NAME", "File Name",
        "MODULE", "Module", "DOC TYPE", "Doc Type", "DOCUMENT TYPE",
        "Transaction Type", "TRANSACTION TYPE", "TYPE", "Type",
    ]

    source_text = " ".join(clean(find_value(row, [field])) for field in source_fields).upper()

    repair_source_words = [
        "REPAIR ORDER", "REPAIR_ORDER", "REPAIR HISTORY", "REPAIR_HISTORY",
        "RO HISTORY", "RO_HISTORY", "REPAIR FILE", "REPAIR_FILE",
        "WORK ORDER", "WORK_ORDER",
    ]

    return any(word in source_text for word in repair_source_words)

def filter_pricing_rows(rows):
    """Remove rows only when they clearly came from a repair-history source file.

    This keeps RP and IN market conditions visible while keeping separate repair
    history from contaminating outright buy/sell guidance.
    """
    clean_rows = []
    removed = 0

    for row in rows:
        if is_repair_source_row(row):
            removed += 1
        else:
            clean_rows.append(row)

    return clean_rows, removed


def truthy(value):
    text = clean(value).strip().lower()
    return text in ["true", "yes", "y", "1", "cancelled", "canceled"]


def row_is_cancelled(row):
    return truthy(find_value(row, ["CANCELLED", "Cancelled", "CANCELED", "Canceled", "Void", "VOID"]))

def get_outgoing_matches(rows, search):
    return rows_for_part(rows, search)

def get_incoming_matches(rows, search, include_ar=False):
    results = []
    for row in rows_for_part(rows, search):
        cond = clean_cond(find_value(row, ["COND", "Condition", "CONDITION"]))
        if cond == "AR" and not include_ar:
            continue
        results.append(row)
    return results

def get_purchase_history_matches(rows, search, include_ar=False):
    results = []
    for row in rows_for_part(rows, search):
        if row_is_cancelled(row):
            continue
        cond = row_condition(row)
        if cond == "AR" and not include_ar:
            continue
        results.append(row)
    return results

def get_sales_history_matches(rows, search, include_ar=False):
    results = []
    for row in rows_for_part(rows, search):
        if row_is_cancelled(row):
            continue
        cond = row_condition(row)
        if cond == "AR" and not include_ar:
            continue
        results.append(row)
    return results

def row_price(row, side):
    names = BUY_PRICE_FIELD_NAMES if side == "buy" else SELL_PRICE_FIELD_NAMES
    return money(find_value(row, names))

def row_condition(row):
    return clean_cond(find_value(row, COND_FIELD_NAMES))

def row_date(row):
    return parse_date(find_value(row, DATE_FIELD_NAMES))

def date_weight(dt):
    """Recent market data matters more because pricing can move fast."""
    if not dt:
        return 0.35
    age_days = max((TODAY - dt).days, 0)
    if age_days <= 90:
        return 1.00
    if age_days <= 180:
        return 0.85
    if age_days <= 365:
        return 0.65
    if age_days <= 730:
        return 0.35
    return 0.15

def weighted_median(values):
    """values is a list of (price, weight)."""
    clean_values = [(float(v), float(w)) for v, w in values if v and v > 1 and w and w > 0]
    if not clean_values:
        return None
    clean_values.sort(key=lambda x: x[0])
    total = sum(w for _, w in clean_values)
    running = 0
    for value, weight in clean_values:
        running += weight
        if running >= total / 2:
            return value
    return clean_values[-1][0]

def newest_date_for_condition(rows, side, cond):
    dates = []
    for row in rows:
        if row_condition(row) != cond:
            continue
        if not row_price(row, side):
            continue
        dt = row_date(row)
        if dt:
            dates.append(dt)
    return max(dates) if dates else None

def median_for_condition(rows, side, cond, recent_only=False):
    prices = []
    for row in rows:
        if row_condition(row) != cond:
            continue
        price = row_price(row, side)
        if not price or price <= 1:
            continue
        dt = row_date(row)
        if recent_only:
            if not dt:
                continue
            if (TODAY - dt).days > 540:
                continue
        prices.append(price)
    if not prices:
        return None
    clean_prices, _ = remove_outliers(prices)
    return median(clean_prices) if clean_prices else None

def apply_cross_condition_trend(guidance, rows, side):
    """Project stale/missing OH using recent RP movement when history supports it.

    Example: old RP 10k and old OH 14k gives OH/RP ratio 1.40.
    If recent RP is 12.5k and no recent OH exists, projected OH is 17.5k.
    """
    notes = []
    if not rows:
        return notes

    old_oh = median_for_condition(rows, side, "OH", recent_only=False)
    old_rp = median_for_condition(rows, side, "RP", recent_only=False)
    recent_rp = median_for_condition(rows, side, "RP", recent_only=True)
    recent_oh = median_for_condition(rows, side, "OH", recent_only=True)
    oh_latest = newest_date_for_condition(rows, side, "OH")
    rp_latest = newest_date_for_condition(rows, side, "RP")

    if old_oh and old_rp and recent_rp and old_rp > 1:
        rp_is_newer = False
        if rp_latest and not oh_latest:
            rp_is_newer = True
        elif rp_latest and oh_latest and (rp_latest - oh_latest).days > 180:
            rp_is_newer = True
        elif recent_oh is None:
            rp_is_newer = True

        if rp_is_newer:
            ratio = old_oh / old_rp
            # Keep ratio realistic so one bad historic point does not explode pricing.
            ratio = max(1.10, min(ratio, 2.25))
            projected_oh = recent_rp * ratio
            current_oh = None
            if "OH" in guidance:
                current_oh = guidance["OH"].get("median") or guidance["OH"].get("start") or guidance["OH"].get("target")

            if current_oh is None or projected_oh > current_oh * 1.08:
                if side == "buy":
                    guidance["OH"] = {
                        "target": projected_oh * 0.90,
                        "max": projected_oh * 1.10,
                        "median": projected_oh,
                        "count": max(guidance.get("OH", {}).get("count", 0), 2),
                        "source": guidance.get("OH", {}).get("source", "condition-adjusted market trend"),
                        "trend_adjusted": True,
                    }
                else:
                    guidance["OH"] = {
                        "start": projected_oh,
                        "low": projected_oh * 0.90,
                        "high": projected_oh * 1.10,
                        "count": max(guidance.get("OH", {}).get("count", 0), 2),
                        "source": guidance.get("OH", {}).get("source", "condition-adjusted market trend"),
                        "trend_adjusted": True,
                    }
                notes.append(f"OH adjusted from recent RP trend using OH/RP ratio {ratio:.2f}x")
    return notes

def transaction_guidance(rows, side, source_name):
    by_cond = defaultdict(list)
    base_source_weight = SOURCE_WEIGHT.get(source_name, 1.0)

    for row in rows:
        if is_repair_source_row(row) or row_is_cancelled(row):
            continue
        cond = row_condition(row)
        price = row_price(row, side)
        if cond in VISIBLE_BUY_CONDITIONS and price and price > 1:
            weight = base_source_weight * date_weight(row_date(row))
            by_cond[cond].append((price, weight))

    guidance = {}
    for cond, weighted_prices in by_cond.items():
        raw_prices = [p for p, _ in weighted_prices]
        clean_prices, removed = remove_outliers(raw_prices)
        allowed = set(clean_prices)
        weighted_prices = [(p, w) for p, w in weighted_prices if p in allowed]
        med = weighted_median(weighted_prices)
        if not med or med <= 1:
            continue

        effective_weight = sum(w for _, w in weighted_prices)
        if side == "buy":
            guidance[cond] = {
                "target": med * 0.90,
                "max": med * 1.10,
                "median": med,
                "count": len(weighted_prices),
                "effective_weight": effective_weight,
                "source": source_name,
            }
        else:
            guidance[cond] = {
                "start": med,
                "low": med * 0.90,
                "high": med * 1.10,
                "count": len(weighted_prices),
                "effective_weight": effective_weight,
                "source": source_name,
            }

    return guidance

def merge_buy_guidance(primary, secondary):
    # Completed PO history is the truth. Incoming quotes are fallback/supporting data.
    merged = dict(primary)
    for cond, info in secondary.items():
        if cond not in merged:
            merged[cond] = info
        else:
            merged[cond]["support_count"] = info.get("count", 0)
            merged[cond]["support_source"] = info.get("source", "incoming quotes")
    return merged

def merge_sell_guidance(primary, secondary):
    # Completed sales history is the truth. Outgoing quotes are fallback/supporting data.
    merged = dict(primary)
    for cond, info in secondary.items():
        if cond not in merged:
            merged[cond] = info
        else:
            merged[cond]["support_count"] = info.get("count", 0)
            merged[cond]["support_source"] = info.get("source", "outgoing quotes")
    return merged

def sell_recommendations(outgoing_matches):
    return transaction_guidance(outgoing_matches, "sell", "outgoing quotes")

def explicit_warranty_days(value):
    text = clean(value).lower()
    if not text or text in ["none", "no", "n/a", "na", "0"]:
        return None

    digits = ""
    for ch in text:
        if ch.isdigit():
            digits += ch
        elif digits:
            break

    if not digits:
        return None

    n = int(digits)

    if "year" in text or "yr" in text:
        return n * 365
    if "month" in text or "mo" in text:
        return n * 30
    if "day" in text or "dy" in text:
        return n

    return n

def tag_based_warranty_remaining(row):
    cond = clean_cond(find_value(row, ["COND", "Condition", "CONDITION"]))
    tag = parse_date(find_value(row, ["TAG DATE", "Tag Date", "Tag", "Cert Date"]))

    if cond == "IN":
        return 30

    if not tag:
        return None

    try:
        if cond == "OH":
            warranty_end = tag.replace(year=tag.year + 1)
        elif cond == "RP":
            warranty_end = tag + timedelta(days=180)
        else:
            return None
    except:
        return None

    remaining = (warranty_end - TODAY).days
    return remaining if remaining > 0 else None

def warranty_days(row):
    explicit = explicit_warranty_days(find_value(row, ["WARRANTY", "Warranty"]))
    if explicit is not None:
        return explicit

    tag_remaining = tag_based_warranty_remaining(row)
    if tag_remaining is not None:
        return tag_remaining

    return 30

def tag_bucket(row):
    tag = parse_date(find_value(row, ["TAG DATE", "Tag Date", "Tag", "Cert Date"]))
    if not tag:
        return 0
    return tag.year * 12 + tag.month

def buy_guidance(incoming_matches):
    return transaction_guidance(incoming_matches, "buy", "incoming quotes")

def condition_medians(incoming_matches):
    return {cond: info["median"] for cond, info in buy_guidance(incoming_matches).items()}

def is_reasonable_offer(row, cond_medians):
    cond = clean_cond(find_value(row, ["COND", "Condition", "CONDITION"]))
    cost = money(find_value(row, ["COST PER UNIT", "Cost Per Unit", "Unit Cost", "Cost", "Price"]))

    if not cost or cost <= 0:
        return False

    med = cond_medians.get(cond)
    if med:
        return cost <= med * 1.25

    return True

def best_buy_score(row):
    cond = clean_cond(find_value(row, ["COND", "Condition", "CONDITION"]))
    cost = money(find_value(row, ["COST PER UNIT", "Cost Per Unit", "Unit Cost", "Cost", "Price"])) or 999999999

    score = 0
    score -= cost / 1000
    score += tag_bucket(row) * 0.6
    score += COND_RANK.get(cond, 0) * 18

    warr = warranty_days(row)
    if warr >= 365:
        score += 50
    elif warr >= 180:
        score += 30
    elif warr >= 90:
        score += 15
    elif warr >= 30:
        score += 5

    return score

def quality_score(row):
    cond = clean_cond(find_value(row, ["COND", "Condition", "CONDITION"]))
    cost = money(find_value(row, ["COST PER UNIT", "Cost Per Unit", "Unit Cost", "Cost", "Price"])) or 999999999

    score = 0
    score += tag_bucket(row) * 1.0
    score += COND_RANK.get(cond, 0) * 25
    score -= cost / 2500

    warr = warranty_days(row)
    if warr >= 365:
        score += 60
    elif warr >= 180:
        score += 40
    elif warr >= 90:
        score += 20
    elif warr >= 30:
        score += 8

    return score

def short_offer(row):
    cond = clean_cond(find_value(row, ["COND", "Condition", "CONDITION"]))
    cost = money(find_value(row, ["COST PER UNIT", "Cost Per Unit", "Unit Cost", "Cost", "Price"])) or 0
    tag = parse_date(find_value(row, ["TAG DATE", "Tag Date", "Tag", "Cert Date"]))
    tag_text = tag.strftime("%Y tag") if tag else "no tag"
    return f"{cond} | {fmt_money(cost)} | {tag_text}"

def pick_buy_options(incoming_matches):
    usable = []

    for row in incoming_matches:
        cond = clean_cond(find_value(row, ["COND", "Condition", "CONDITION"]))
        cost = money(find_value(row, ["COST PER UNIT", "Cost Per Unit", "Unit Cost", "Cost", "Price"]))

        if cond in VISIBLE_BUY_CONDITIONS and cost and cost > 1:
            usable.append(row)

    if not usable:
        return None, None

    medians = condition_medians(incoming_matches)
    reasonable = [row for row in usable if is_reasonable_offer(row, medians)]

    if not reasonable:
        reasonable = usable

    best = max(reasonable, key=best_buy_score)
    second_pool = [row for row in reasonable if row != best]
    second = max(second_pool, key=quality_score) if second_pool else None

    return best, second

def best_buy_reason(best, second=None):
    best_cond = clean_cond(find_value(best, ["COND", "Condition", "CONDITION"]))
    best_cost = money(find_value(best, ["COST PER UNIT", "Cost Per Unit", "Unit Cost", "Cost", "Price"])) or 0

    if second:
        second_cost = money(find_value(second, ["COST PER UNIT", "Cost Per Unit", "Unit Cost", "Cost", "Price"])) or 0
        if second_cost and best_cost < second_cost * 0.75:
            return "Best overall value with much lower acquisition cost."
        if best_cond == "OH":
            return "Strong condition with good value and acceptable tag position."
        if best_cond == "RP":
            return "Good tag position and usable condition at the best practical value."

    if best_cond == "OH":
        return "Strong overall value with good condition."
    if best_cond == "RP":
        return "Good buying option based on tag, condition, and price."
    if best_cond == "NE":
        return "Highest condition quality with strong resale position."
    return "Best balance of tag, condition, and price."

def second_option_reason(second, best):
    second_cond = clean_cond(find_value(second, ["COND", "Condition", "CONDITION"]))
    best_cond = clean_cond(find_value(best, ["COND", "Condition", "CONDITION"]))
    second_cost = money(find_value(second, ["COST PER UNIT", "Cost Per Unit", "Unit Cost", "Cost", "Price"])) or 0
    best_cost = money(find_value(best, ["COST PER UNIT", "Cost Per Unit", "Unit Cost", "Cost", "Price"])) or 0

    if tag_bucket(second) > tag_bucket(best):
        return "Newer tag option, but review cost before buying."
    if COND_RANK.get(second_cond, 0) > COND_RANK.get(best_cond, 0):
        return "Better condition option, but not the best overall value."
    if second_cost > best_cost:
        return "Higher cost alternate option if first choice is unavailable."
    return "Alternate buying option worth evaluating."

def listing_qty(value):
    text = clean(value).upper()
    if text in ["", "RQST", "RFQ", "CALL"]:
        return 0
    try:
        return int(float(text))
    except:
        return 0

def vendor_qty_for_part(vendor, part, include_ar=False):
    total = 0
    best = 0
    by_cond = defaultdict(int)

    for listing in vendor["listings"]:
        if listing["part"] != part.upper():
            continue

        cond = listing["cond"]
        if cond == "AR" and not include_ar:
            continue

        qty = listing_qty(listing.get("qty", ""))
        total += qty
        best = max(best, qty)
        by_cond[cond] += qty

    return total, best, dict(by_cond)

def vendor_has_aog(vendor):
    text = " ".join(vendor.get("contacts", [])).upper()
    return "AOG" in text or "24/7" in text or "24 HOUR" in text or "AFTER HOURS" in text

def action_label(idx, info):
    if idx == 1:
        return "CALL FIRST"
    if idx <= 3:
        return "CALL NEXT"
    if info["section"] == "old":
        return "LOW PRIORITY - STALE LISTING"
    return "BACKUP"

def call_reasons(info):
    reasons = []

    if info.get("relationship_note"):
        reasons.append(info["relationship_note"])

    if info["part_po"] > 0:
        reasons.append(f"bought this P/N {info['part_po']} time(s)")
    elif info["total_po"] > 0:
        if info.get("known_company_fallback"):
            reasons.append(f"known supplier fallback: {info['total_po']} prior PO(s)")
        else:
            reasons.append(f"{info['total_po']} prior PO(s)")

    if info["qty_total"] > 0:
        reasons.append(f"{info['qty_total']} total listed")

    if info["conds"]:
        reasons.append("conditions " + "/".join(info["conds"]))

    if info["section"] != "unknown":
        reasons.append(f"ILS {info['section']}")

    if info["has_aog"]:
        reasons.append("AOG/contact info")

    if not reasons:
        reasons.append("usable ILS listing")

    return "; ".join(reasons)

def ils_vendor_score(vendor, part, po_history, include_ar=False, exact_part_history_exists=False):
    listings = [l for l in vendor["listings"] if l["part"] == part.upper()]

    if not include_ar:
        listings = [l for l in listings if l["cond"] != "AR"]

    if not listings:
        return None

    conds = sorted(set(l["cond"] for l in listings), key=lambda c: COND_RANK.get(c, 0), reverse=True)
    best_cond_rank = max(COND_RANK.get(c, 0) for c in conds)
    section_score = SECTION_RANK.get(vendor["section"], 0)
    qty_total, qty_best, qty_by_cond = vendor_qty_for_part(vendor, part, include_ar=include_ar)
    has_aog = vendor_has_aog(vendor)

    vendor_history = po_history.get(vendor["vendor"], {"total_po_count": 0, "part_po_count": defaultdict(int)})
    total_po = vendor_history["total_po_count"]
    part_po = vendor_history["part_po_count"][part.upper()]

    relationship_tier, relationship_note = company_relationship(vendor["vendor"])

    # Primary objective: who to call first.
    #
    # New rule: purchasing relationship is the strongest signal. A vendor with
    # many historical POs should generally rank above a vendor we rarely/never
    # buy from, even if both show stock. Exact same-P/N history still matters,
    # but total PO relationship now carries the highest normal weight.
    score = 0

    # Total PO history = strongest normal scoring factor.
    score += min(total_po, 500) * 18

    # Exact same-part purchase history = strong extra boost.
    score += min(part_po, 25) * 110

    # If nobody has exact same-part history, give known suppliers an even
    # stronger fallback bump so familiar companies beat unknown vendors.
    if not exact_part_history_exists and total_po > 0:
        score += min(total_po, 500) * 7

    # Preferred family companies get first-call priority when they show stock.
    # Final sorting also uses relationship_tier, so these appear before outside
    # vendors: our own company first, then sister companies.
    if relationship_tier == 0:
        score += 100000
    elif relationship_tier == 1:
        score += 90000

    score += min(qty_total, 30) * 4
    score += min(qty_best, 10) * 3
    score += section_score * 22
    score += best_cond_rank * 14

    if len(conds) >= 2:
        score += 8
    if vendor["contacts"]:
        score += 6
    if has_aog:
        score += 8
    if vendor["section"] == "old":
        score -= 40
    if conds == ["AR"]:
        score -= 60

    return {
        "score": score,
        "conds": conds,
        "section": vendor["section"],
        "contacts": vendor["contacts"],
        "total_po": total_po,
        "part_po": part_po,
        "qty_total": qty_total,
        "qty_best": qty_best,
        "qty_by_cond": qty_by_cond,
        "has_aog": has_aog,
        "relationship_tier": relationship_tier,
        "relationship_note": relationship_note,
        "known_company_fallback": (part_po == 0 and total_po > 0 and not exact_part_history_exists),
    }

def rank_ils_vendors(ils_vendors, part, po_history, include_ar=False):
    ranked = []

    # Check whether any listed ILS vendor has exact-part PO history.
    # If not, ranking leans more on vendors Support Air has bought from before.
    exact_part_history_exists = False
    for vendor in ils_vendors:
        if not any(l["part"] == part.upper() and (include_ar or l["cond"] != "AR") for l in vendor["listings"]):
            continue
        vendor_history = po_history.get(vendor["vendor"], {"part_po_count": defaultdict(int)})
        if vendor_history["part_po_count"][part.upper()] > 0:
            exact_part_history_exists = True
            break

    for vendor in ils_vendors:
        info = ils_vendor_score(
            vendor,
            part,
            po_history,
            include_ar=include_ar,
            exact_part_history_exists=exact_part_history_exists,
        )
        if info is not None:
            ranked.append((info["score"], vendor, info))

    # Preferred company order:
    #   0 = our own company
    #   1 = sister/preferred company
    #   2 = normal outside vendor
    # Then sort by the intelligence score.
    ranked.sort(key=lambda x: (x[2].get("relationship_tier", 2), -x[0]))
    return ranked


def get_inventory_matches(rows, search, include_ar=False):
    """Return all inventory rows for this part number as historical transaction intelligence.

    No on-hand filtering is applied. Every row is treated as an internal history record.
    """
    results = []
    search = search.upper()
    for row in rows:
        pn = clean(find_value(row, PART_FIELD_NAMES)).upper()
        if pn != search:
            continue
        cond = row_condition(row)
        if cond == "AR" and not include_ar:
            continue
        results.append(row)
    return results

def summarize_inventory(inventory_matches):
    """Build internal history intelligence from inventory rows.

    Fields used per spec:
      Date, P/N, Description, Total Cost, SRP, EXCH SRP,
      Purchased From, Tagged By, Sale Price

    No on-hand quantity is inferred. qty field is intentionally omitted.
    """
    summary = defaultdict(lambda: {
        "total_costs": [],
        "srps": [],
        "exch_srps": [],
        "sale_prices": [],
        "vendors": [],        # Purchased From -> vendor reliability
        "tagged_by": [],      # Tagged By -> repair capability
        "dates": [],
        "descriptions": [],
    })
    for row in inventory_matches:
        cond = row_condition(row)
        if not cond:
            cond = "UNK"
        info = summary[cond]

        total_cost = money(find_value(row, INVENTORY_COST_FIELD_NAMES))
        srp = money(find_value(row, INVENTORY_SRP_FIELD_NAMES))
        exch_srp = money(find_value(row, INVENTORY_EXCH_SRP_FIELD_NAMES))
        sale_price = money(find_value(row, INVENTORY_SALE_PRICE_FIELD_NAMES))
        vendor = normalize_vendor(find_value(row, INVENTORY_VENDOR_FIELD_NAMES))
        tagged_by = clean(find_value(row, INVENTORY_TAG_BY_FIELD_NAMES))
        dt = parse_date(find_value(row, DATE_FIELD_NAMES))
        desc = clean(find_value(row, ["DESCRIPTION", "Description", "DESC"]))

        if total_cost and total_cost > 0:
            info["total_costs"].append(total_cost)
        if srp and srp > 0:
            info["srps"].append(srp)
        if exch_srp and exch_srp > 0:
            info["exch_srps"].append(exch_srp)
        if sale_price and sale_price > 0:
            info["sale_prices"].append(sale_price)
        if vendor:
            info["vendors"].append(vendor)
        if tagged_by:
            info["tagged_by"].append(tagged_by)
        if dt:
            info["dates"].append(dt)
        if desc and desc not in info["descriptions"]:
            info["descriptions"].append(desc)

    return summary

def inventory_sell_floor(summary, cond):
    """Derive a sell floor from internal history.

    Priority:
      1. Completed sale price history (most authoritative)
      2. Total cost history with margin floor
      3. SRP as reference only (lowest priority)
    Do NOT let old SRP override recent completed sale or cost data.
    """
    info = summary.get(cond)
    if not info:
        return None

    # 1. Completed sale history
    if info["sale_prices"]:
        return median(info["sale_prices"])

    # 2. Acquisition cost + margin floor
    if info["total_costs"]:
        cost_floor = median(info["total_costs"]) * 1.35
        # If SRP is higher, use it, but only as a soft cap — not an override
        if info["srps"]:
            srp_ref = median(info["srps"])
            return max(cost_floor, srp_ref * 0.85)
        return cost_floor

    # 3. SRP reference only — weakest signal
    if info["srps"]:
        return median(info["srps"]) * 0.80

    return None

def show_inventory_snapshot(inventory_rows, part, include_ar=False, details=False):
    """Display internal transaction/inventory intelligence.

    Does NOT say 'X on hand' — there is no confirmed on-hand quantity field.
    Uses: Total Cost, SRP, EXCH SRP, Sale Price, Purchased From, Tagged By.
    """
    matches = get_inventory_matches(inventory_rows, part, include_ar=include_ar)
    print("\n===== INTERNAL HISTORY =====")
    if not matches:
        print("No internal history found for this part.")
        return {}

    summary = summarize_inventory(matches)

    for cond in ["NE", "OH", "RP", "IN", "SV", "NS", "AR", "UNK"]:
        if cond not in summary:
            continue
        info = summary[cond]
        record_count = len([
            1 for row in matches
            if (row_condition(row) or "UNK") == cond
        ])

        print(f"\n{cond} | {record_count} record(s) found")

        # Purchase history line
        purchase_parts = []
        if info["total_costs"]:
            purchase_parts.append(f"total cost {fmt_money(median(info['total_costs']))}")
        if info["srps"]:
            purchase_parts.append(f"SRP {fmt_money(median(info['srps']))}")
        if info["exch_srps"]:
            purchase_parts.append(f"EXCH SRP {fmt_money(median(info['exch_srps']))}")
        if info["vendors"]:
            unique_vendors = list(dict.fromkeys(info["vendors"]))[:3]
            purchase_parts.append("purchased from " + " / ".join(unique_vendors))
        if purchase_parts:
            print("  Purchase history: " + " | ".join(purchase_parts))

        # Tagged by (repair capability intelligence)
        if info["tagged_by"]:
            unique_tags = list(dict.fromkeys(info["tagged_by"]))[:3]
            print("  Repair/tag source: tagged by " + " / ".join(unique_tags))

        # Sale history
        if info["sale_prices"]:
            sale_vals = sorted(info["sale_prices"])
            if len(sale_vals) == 1:
                print(f"  Prior sale: {fmt_money(sale_vals[0])}")
            else:
                print(f"  Prior sales: {fmt_money(sale_vals[0])} – {fmt_money(sale_vals[-1])} | median {fmt_money(median(sale_vals))}")

        # Dates
        if details and info["dates"]:
            newest = max(info["dates"])
            oldest = min(info["dates"])
            if newest == oldest:
                print(f"  Date: {newest.strftime('%Y-%m-%d')}")
            else:
                print(f"  Date range: {oldest.strftime('%Y-%m-%d')} – {newest.strftime('%Y-%m-%d')}")

    return summary

def show_market_snapshot(ils_vendors, part, include_ar=False):
    cond_counts = Counter()
    vendor_count = 0

    for vendor in ils_vendors:
        vendor_has_part = False
        for listing in vendor["listings"]:
            if listing["part"] != part.upper():
                continue
            cond = listing["cond"]
            if cond == "AR" and not include_ar:
                continue
            qty = listing_qty(listing.get("qty", ""))
            cond_counts[cond] += qty if qty > 0 else 1
            vendor_has_part = True
        if vendor_has_part:
            vendor_count += 1

    print("\n===== MARKET SNAPSHOT =====")
    if not cond_counts:
        print("No usable ILS availability found.")
        return

    print(f"ILS vendor groups with usable listings: {vendor_count}")
    for cond in ["NE", "OH", "RP", "IN", "SV", "NS", "AR"]:
        if cond in cond_counts:
            print(f"{cond}: {cond_counts[cond]} listed")

def show_call_list(ils_vendors, po_history, part, details=False, include_ar=False):
    print("\n===== WHO TO CALL FIRST =====")

    if not ils_vendors:
        print("No ILS file loaded.")
        return

    ranked = rank_ils_vendors(ils_vendors, part, po_history, include_ar=include_ar)

    if not ranked:
        print("No usable ILS vendor listings found.")
        print("Check that the matching ILS .txt file is in C:\\Quote_Tool\\DATA and that the part number is exact.")
        return

    for idx, (score, vendor, info) in enumerate(ranked[:8], start=1):
        print(f"{idx}. {vendor['vendor']} | {action_label(idx, info)}")
        print(f"   Why: {call_reasons(info)}")

        if details:
            print(f"   Score: {round(score, 1)}")
            if info["qty_by_cond"]:
                qty_text = ", ".join(f"{cond} {qty}" for cond, qty in sorted(info["qty_by_cond"].items(), key=lambda x: COND_RANK.get(x[0], 0), reverse=True))
                print(f"   Qty by condition: {qty_text}")
            if info["contacts"]:
                print(f"   Contact: {info['contacts'][0]}")

def apply_condition_sanity(buy, sell):
    """Keep directional condition values logical. RP/IN/SV should not exceed OH.

    This does not delete any condition. It only caps distorted guidance when old history
    makes a lower condition price above OH. Details mode can show that a sanity cap was used.
    """
    adjusted = []

    if "OH" in buy:
        oh = buy["OH"]
        for cond, pct in CONDITION_SANITY_CAPS.items():
            info = buy.get(cond)
            if not info:
                continue
            old_target = info.get("target")
            old_max = info.get("max")
            if old_target is not None and old_target > oh["target"] * pct:
                info["target"] = oh["target"] * pct
                info["sanity_adjusted"] = True
            if old_max is not None and old_max > oh["max"] * pct:
                info["max"] = oh["max"] * pct
                info["sanity_adjusted"] = True
            if info.get("sanity_adjusted"):
                adjusted.append(f"{cond} buy capped below OH")

    if "OH" in sell:
        oh = sell["OH"]
        oh_start = oh.get("start", 0)
        for cond, pct in CONDITION_SANITY_CAPS.items():
            info = sell.get(cond)
            if not info or not oh_start:
                continue
            if info.get("start", 0) > oh_start * pct:
                info["start"] = oh_start * pct
                info["low"] = info["start"] * 0.90
                info["high"] = info["start"] * 1.10
                info["sanity_adjusted"] = True
                adjusted.append(f"{cond} sell capped below OH")

    return adjusted

def build_price_debug_summary(purchase_matches_raw, purchase_matches, purchase_removed, incoming_matches_raw, incoming_matches, incoming_removed, sales_matches_raw, sales_matches, sales_removed, outgoing_matches_raw, outgoing_matches, outgoing_removed, sanity_notes):
    print("\nPricing detail:")
    print(f"Completed PO records: {len(purchase_matches_raw)} total | {len(purchase_matches)} used | {purchase_removed} repair-source rows ignored")
    print(f"Incoming quote records: {len(incoming_matches_raw)} total | {len(incoming_matches)} used | {incoming_removed} repair-source rows ignored")
    print(f"Completed sales records: {len(sales_matches_raw)} total | {len(sales_matches)} used | {sales_removed} repair-source rows ignored")
    print(f"Outgoing quote records: {len(outgoing_matches_raw)} total | {len(outgoing_matches)} used | {outgoing_removed} repair-source rows ignored")
    if sanity_notes:
        print("Condition sanity adjustments: " + "; ".join(dict.fromkeys(sanity_notes)))

def show_buy_quote(part, outgoing_rows, incoming_rows, po_rows=None, sales_rows=None, inventory_rows=None, details=False):
    po_rows = po_rows or []
    sales_rows = sales_rows or []
    inventory_rows = inventory_rows or []
    inventory_matches = get_inventory_matches(inventory_rows, part, include_ar=False)
    inventory_summary = summarize_inventory(inventory_matches)

    outgoing_matches_raw = get_outgoing_matches(outgoing_rows, part)
    incoming_matches_raw = get_incoming_matches(incoming_rows, part, include_ar=False)
    purchase_matches_raw = get_purchase_history_matches(po_rows, part, include_ar=False)
    sales_matches_raw = get_sales_history_matches(sales_rows, part, include_ar=False)

    # Pricing should be directional outright market guidance.
    # Completed transactions get priority; quote history is fallback/support.
    outgoing_matches, outgoing_removed = filter_pricing_rows(outgoing_matches_raw)
    incoming_matches, incoming_removed = filter_pricing_rows(incoming_matches_raw)
    purchase_matches, purchase_removed = filter_pricing_rows(purchase_matches_raw)
    sales_matches, sales_removed = filter_pricing_rows(sales_matches_raw)

    buy_from_po = transaction_guidance(purchase_matches, "buy", "completed PO history")
    buy_from_quotes = buy_guidance(incoming_matches)
    buy = merge_buy_guidance(buy_from_po, buy_from_quotes)

    sell_from_sales = transaction_guidance(sales_matches, "sell", "completed sales history")
    sell_from_quotes = sell_recommendations(outgoing_matches)
    sell = merge_sell_guidance(sell_from_sales, sell_from_quotes)

    buy_trend_notes = apply_cross_condition_trend(buy, purchase_matches + incoming_matches, "buy")
    sell_trend_notes = apply_cross_condition_trend(sell, sales_matches + outgoing_matches, "sell")
    sanity_notes = apply_condition_sanity(buy, sell)

    print("\n===== PRICING GUIDANCE =====")
    if details:
        build_price_debug_summary(
            purchase_matches_raw, purchase_matches, purchase_removed,
            incoming_matches_raw, incoming_matches, incoming_removed,
            sales_matches_raw, sales_matches, sales_removed,
            outgoing_matches_raw, outgoing_matches, outgoing_removed,
            sanity_notes + buy_trend_notes + sell_trend_notes,
        )
        if buy_trend_notes or sell_trend_notes:
            print("Trend adjustments: " + "; ".join(dict.fromkeys(buy_trend_notes + sell_trend_notes)))

    print("\nBUY SIDE / ACQUISITION GUIDANCE")
    if buy:
        for cond in VISIBLE_BUY_CONDITIONS:
            if cond in buy:
                info = buy[cond]
                detail = ""
                if details:
                    support = ""
                    if info.get("support_count"):
                        support = f"; supported by {info['support_count']} {info.get('support_source', 'quote')}"
                    sanity = "; sanity capped below OH" if info.get("sanity_adjusted") else ""
                    trend = "; trend adjusted" if info.get("trend_adjusted") else ""
                    detail = f" | {info.get('source', 'pricing history')}{support}{sanity}{trend}"
                print(
                    f"{cond}: Expected range "
                    f"{fmt_money(info['target'])} - {fmt_money(info['max'])} "
                    f"| {confidence(info['count'])}{detail}"
                )
    else:
        print("No usable buy-side pricing found.")

    print("\nSELL SIDE / QUOTE GUIDANCE")
    printed_sell = False
    adjusted_any = False
    prepared_sell = {}

    for cond in ["OH", "RP", "IN", "SV", "NE"]:
        raw_info = sell.get(cond)
        buy_info = buy.get(cond)

        if raw_info:
            quote_value = raw_info["start"]
            label = "Quote around"
            conf = confidence(raw_info["count"])
            source = raw_info.get("source", "pricing history")

            if buy_info:
                sell_floor = buy_info["max"] * MIN_SELL_MARKUP_OVER_BUY_MAX
                if quote_value < sell_floor:
                    quote_value = sell_floor
                    label = "Quote floor around"
                    source += "; raised because buy market is higher than sell history"
                    adjusted_any = True

            inv_floor = inventory_sell_floor(inventory_summary, cond)
            if inv_floor and quote_value < inv_floor:
                quote_value = inv_floor
                label = "Quote floor around"
                source += "; raised by internal history floor"
                adjusted_any = True

            prepared_sell[cond] = {
                "value": quote_value,
                "label": label,
                "conf": conf,
                "source": source,
                "support_count": raw_info.get("support_count"),
                "support_source": raw_info.get("support_source", "quote"),
                "sanity_adjusted": raw_info.get("sanity_adjusted"),
                "trend_adjusted": raw_info.get("trend_adjusted"),
            }

        elif buy_info:
            value = buy_info["max"] * MIN_SELL_MARKUP_OVER_BUY_MAX
            inv_floor = inventory_sell_floor(inventory_summary, cond)
            if inv_floor and value < inv_floor:
                value = inv_floor
            prepared_sell[cond] = {
                "value": value,
                "label": "Quote floor around",
                "conf": confidence(buy_info["count"]),
                "source": "based on buy market",
                "support_count": 0,
                "support_source": "",
                "sanity_adjusted": False,
            }
            adjusted_any = True

    # Final sanity pass after buy-market floors. RP/IN/SV should not print above OH.
    if "OH" in prepared_sell:
        oh_value = prepared_sell["OH"]["value"]
        for cond, pct in CONDITION_SANITY_CAPS.items():
            if cond in prepared_sell and prepared_sell[cond]["value"] > oh_value * pct:
                prepared_sell[cond]["value"] = oh_value * pct
                prepared_sell[cond]["label"] = "Quote around"
                prepared_sell[cond]["sanity_adjusted"] = True
                adjusted_any = True

    for cond in ["OH", "RP", "IN", "SV", "NE"]:
        info = prepared_sell.get(cond)
        if not info:
            continue
        detail = ""
        if details:
            support = ""
            if info.get("support_count"):
                support = f"; supported by {info['support_count']} {info.get('support_source', 'quote')}"
            sanity = "; sanity capped below OH" if info.get("sanity_adjusted") else ""
            trend = "; trend adjusted" if info.get("trend_adjusted") else ""
            detail = f" | {info['source']}{support}{sanity}{trend}"
        print(f"{cond}: {info['label']} {fmt_money(info['value'])} | {info['conf']}{detail}")
        printed_sell = True

    if not printed_sell:
        print("No usable sell-side pricing found.")

    if details:
        print(
            "\nTrend note: pricing is directional only. "
            "Completed PO/SO history is prioritized over quote history when available. "
            "RP and IN remain valid conditions; only rows clearly marked as repair-history/source data are ignored."
        )

def show_top_parts(outgoing_rows):
    counts = Counter()
    for row in outgoing_rows:
        pn = clean(find_value(row, ["Part #", "PART #", "Part Number", "PART NUMBER", "P/N", "PN"]))
        if pn:
            counts[pn] += 1

    print("\nMOST QUOTED PARTS")
    for pn, count in counts.most_common(10):
        print(f"{pn}: {count}")

def main():
    print("Loading market intelligence files...")

    outgoing_file = find_data_file(OUTGOING_FILE_CANDIDATES)
    incoming_file = find_data_file(INCOMING_FILE_CANDIDATES)
    po_file = find_data_file(PO_FILE_CANDIDATES)

    outgoing_rows = read_excel_file(outgoing_file) if outgoing_file else []
    incoming_rows = read_excel_file(incoming_file) if incoming_file else []
    po_rows = read_excel_file(po_file) if po_file else []
    sales_rows, sales_file = read_first_existing_excel(SALES_HISTORY_CANDIDATES)
    inventory_file = find_data_file(INVENTORY_FILE_CANDIDATES)
    inventory_rows = []  # loaded by part on demand
    active_ils_files = find_all_ils_files()
    ils_vendors = []
    for ils_file in active_ils_files:
        ils_vendors.extend(parse_ils_file(ils_file))

    po_history = load_po_history(po_rows)
    outgoing_index = build_part_index(outgoing_rows, ["Part #", "PART #", "Part Number", "PART NUMBER", "P/N", "PN"])
    incoming_index = build_part_index(incoming_rows, ["PART #", "Part #", "Part Number", "PART NUMBER", "P/N", "PN"])
    po_index = build_part_index(po_rows, PART_FIELD_NAMES)
    sales_index = build_part_index(sales_rows, PART_FIELD_NAMES)

    print(f"Loaded {len(outgoing_rows)} outgoing quote rows" + (f" from {outgoing_file.name}." if outgoing_file else ". Add outgoing_quotes.xlsx to DATA."))
    print(f"Loaded {len(incoming_rows)} incoming quote rows" + (f" from {incoming_file.name}." if incoming_file else ". Add incoming_quotes.xlsx to DATA."))
    if active_ils_files:
        print(f"Loaded {len(ils_vendors)} ILS vendor groups from {len(active_ils_files)} ILS text file(s):")
        for ils_file in active_ils_files:
            print(f"  - {ils_file.name}")
    else:
        print("Loaded 0 ILS vendor groups. No .txt ILS files found in DATA.")
    print(f"Loaded PO history for {len(po_history)} vendor(s)" + (f" from {po_file.name}." if po_file else ". Add purchase_orders.xlsx to DATA."))
    if sales_file:
        print(f"Loaded {len(sales_rows)} completed sales row(s) from {sales_file.name}.")
    else:
        print("Loaded 0 completed sales rows. Add sales_orders.xlsx, so.xlsx, sales_history.xlsx, completed_sales.xlsx, or invoices.xlsx to DATA to use actual sales history.")
    if inventory_file:
        print(f"Inventory file ready: {inventory_file.name}. Rows load by part number on demand.")
    else:
        print("Loaded 0 inventory rows. Add inventory.xlsx to DATA to use internal stock awareness.")
    print("Market quote/buy tool ready.")
    print("Commands: top, details, ar, exit")

    details = False
    include_ar = False

    while True:
        part = input("\nEnter exact part number or command: ").strip()

        if part.lower() in ["exit", "quit", "q"]:
            break

        if part.lower() == "top":
            show_top_parts(outgoing_rows)
            continue

        if part.lower() == "details":
            details = not details
            print(f"Details mode: {'ON' if details else 'OFF'}")
            continue

        if part.lower() == "ar":
            include_ar = not include_ar
            print(f"AR mode: {'ON' if include_ar else 'OFF'}")
            continue

        print(f"\nPART: {part.upper()}")
        inventory_rows_for_part = read_excel_file_for_part(inventory_file, part) if inventory_file else []
        show_inventory_snapshot(inventory_rows_for_part, part, include_ar=include_ar, details=details)
        show_call_list(ils_vendors, po_history, part, details=details, include_ar=include_ar)
        show_market_snapshot(ils_vendors, part, include_ar=include_ar)
        show_buy_quote(part, outgoing_index, incoming_index, po_rows=po_index, sales_rows=sales_index, inventory_rows=inventory_rows_for_part, details=details)

if __name__ == "__main__":
    main()
